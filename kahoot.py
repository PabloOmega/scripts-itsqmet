import pandas as pd
from openpyxl import load_workbook

# === 1. Cargar la plantilla de Kahoot ===
template_path = "KahootQuizTemplate.xlsx"   # Ruta a tu plantilla
wb = load_workbook(template_path)
ws = wb.active

# === 2. Definir las preguntas y respuestas ===
# Formato: (Pregunta, Respuesta 1, Respuesta 2, Respuesta 3, Respuesta 4, Tiempo, Respuesta Correcta)
kahoot_questions = [
    ("¿Cuál es la diferencia principal entre MVC y MVT en Django?",
     "En MVT el controlador está integrado en la vista",
     "En MVT no existe la vista",
     "En MVC no hay separación de lógica y presentación",
     "En MVC los templates son obligatorios",
     20, 1),

    ("¿Qué componente de MVT define la estructura de la base de datos y provee métodos para interactuar con ella?",
     "Vista",
     "Modelo",
     "Template",
     "Controlador",
     20, 2),

    ("En MVT, ¿qué elemento contiene los archivos HTML con sintaxis especial para mostrar datos dinámicos?",
     "Modelo",
     "Vista",
     "Template",
     "Framework",
     20, 3),

    ("En el flujo de MVT, después de que la Vista obtiene datos del modelo, ¿qué hace a continuación?",
     "Renderiza directamente la interfaz",
     "Envía los datos a un Template",
     "Devuelve la información al navegador",
     "Solicita datos a la base de datos",
     20, 2),

    ("¿Qué lenguaje se utiliza en Django para incluir lógica mínima dentro de los templates?",
     "Python puro utilizando llaves",
     "Django Template Language (DTL)",
     "JavaScript",
     "Jinja",
     20, 2),

    ("¿Cuál es una de las ventajas clave de MVT en Django?",
     "Facilita el uso de memoria compartida",
     "Elimina la necesidad de modelos de datos",
     "Menos código repetitivo gracias a la automatización",
     "No requiere bases de datos para funcionar",
     20, 3),

    ("En Django, las vistas pueden ser implementadas como:",
     "Funciones o clases",
     "Solo archivos HTML",
     "Scripts en JavaScript",
     "Tablas en la base de datos",
     20, 1),

    ("¿Qué rol cumple el Template en MVT comparado con MVC?",
     "Sustituye al controlador",
     "Actúa como la capa de presentación",
     "Gestiona la lógica de negocio",
     "Almacena los datos en la base de datos",
     20, 2),

    ("¿Qué característica de Django ORM se relaciona directamente con el Modelo de MVT?",
     "Permitir el mapeo de URLs",
     "Automatizar la autenticación de usuarios",
     "Representar tablas de base de datos mediante clases",
     "Renderizar datos en HTML",
     20, 3),

    ("Si un usuario accede a una URL en una aplicación Django, ¿qué componente procesa primero la solicitud?",
     "Modelo",
     "Vista",
     "Template",
     "Framework (enrutador de Django)",
     20, 4),
]




# === 3. Insertar las preguntas en la plantilla ===
# En la plantilla oficial, las preguntas empiezan desde la fila 8
start_row = 9
for i, q in enumerate(kahoot_questions, start=start_row):
    ws.cell(row=i, column=2, value=q[0])  # Pregunta
    ws.cell(row=i, column=3, value=q[1])  # Respuesta 1
    ws.cell(row=i, column=4, value=q[2])  # Respuesta 2
    ws.cell(row=i, column=5, value=q[3])  # Respuesta 3
    ws.cell(row=i, column=6, value=q[4])  # Respuesta 4
    ws.cell(row=i, column=7, value=q[5])  # Tiempo (segundos)
    ws.cell(row=i, column=8, value=q[6])  # Respuesta correcta

# === 4. Guardar el archivo final ===
output_path = "Kahoot_Calidad_Software_Plantilla.xlsx"
wb.save(output_path)

print("✅ Archivo generado en:", output_path)
