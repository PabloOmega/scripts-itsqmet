from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openai import OpenAI
import time
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from vinculacion_types import ActividadRespuesta

client = OpenAI()

template_path = "Vinculacion.xlsx"   # Ruta a tu plantilla
wb = load_workbook(template_path)
ws = wb.active

def obtener_respuesta(prompt):
    response = client.chat.completions.parse(
        # model="gpt-4o-2024-08-06",
        model="gpt-5-mini",
        messages=[
            {"role": "user", "content": prompt}
        ],
        response_format=ActividadRespuesta,
    )
    return response.choices[0].message.parsed

def obtener_actividad_objetivo(actividades):
    response = obtener_respuesta(f"""
                                Realiza un objetivo y una descripción de las actividades que te voy a enviar para 
                                un proyecto de vinculación con la sociedad. El objetivo debe ser claro y conciso.
                                La descripción debe ser corta. 
                                Te mando las actividades: {actividades}
                                Te envío un poco de descripción del proyecto realizado en el Capítulo 4 y 5:
                                Este proyecto tiene como eje central la puesta en marcha del sistema GLPI dentro del 
                                Instituto Nacional de Meteorología e Hidrología del Ecuador (INAMHI), con el propósito de 
                                optimizar el manejo de incidencias técnicas, el control del inventario de equipos tecnológicos y la 
                                gestión documental, utilizando una plataforma de software libre. 
                                La propuesta surge a partir de un interés académico, profesional y social, al detectar la 
                                necesidad de implementar soluciones digitales efectivas en instituciones públicas. Desde el ámbito 
                                académico, representa una oportunidad para aplicar los conocimientos adquiridos en la carrera de 
                                Desarrollo de Software. A nivel profesional, permite fortalecer las habilidades técnicas del equipo 
                                ejecutor, mientras que, desde el enfoque social, busca beneficiar directamente al INAMHI al 
                                mejorar sus procesos internos y operativos. 
                                El objetivo principal de este trabajo es ofrecer una herramienta funcional que contribuya a 
                                sistematizar la atención de solicitudes de soporte, organizar adecuadamente los recursos 
                                tecnológicos y disminuir los tiempos de respuesta, mejorando así la eficiencia operativa y 
                                facilitando decisiones informadas basadas en datos concretos. 
                                La metodología aplicada se basa en un enfoque cualitativo, recurriendo a entrevistas 
                                semiestructuradas para recolectar información relevante sobre la situación actual del área técnica 
                                1 
                                del INAMHI. Esta estrategia permitió identificar con mayor precisión los requerimientos 
                                necesarios para una implementación efectiva del sistema. 
                                El informe se organiza en seis capítulos: el Capítulo 1 aborda el diagnóstico de 
                                necesidades, el trabajo de campo y la línea base; el Capítulo 2 desarrolla los objetivos, productos 
                                esperados, indicadores y estrategias propuestas; el Capítulo 3 plantea el problema, los 
                                antecedentes, la justificación, el tipo de proyecto, el área de vinculación y su localización. El 
                                Capítulo 4 expone la fundamentación científica, presentando los enfoques teóricos y conceptuales 
                                que respaldan el desarrollo de la propuesta y su pertinencia en el contexto de intervención, el 
                                Capítulo 5 detalla la ejecución y puesta en marcha de la propuesta, describiendo las actividades 
                                realizadas, los recursos utilizados y los resultados alcanzados, y por último en el Capítulo 6 se 
                                evalúa la funcionalidad y resultados del proyecto  para verificar si el proyecto cumple con los 
                                requisitos obtenidos durante toda la etapa.  Y para finalizar Cada sección ha sido estructurada para 
                                facilitar la comprensión del propósito y desarrollo de este trabajo de vinculación con la sociedad. 
                                 """)
    return response.actividades

def esperar_elemento(driver, xpath, type=By.XPATH, timeout=20):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((type, xpath))
        )
    except Exception as e:
        print(f"Error al esperar el elemento: {xpath}", e)
        driver.quit()

def obtener_duracion(hora_inicio, hora_fin):
    hora_inicio = datetime.strptime(hora_inicio, "%H:%M")
    hora_fin = datetime.strptime(hora_fin, "%H:%M")
    duracion = hora_fin - hora_inicio
    return duracion

reuniones = [
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/EetuzkFEcwhLhZPbNOu6pNcB-cwNhVa6nW_0ao0zCZYDuA?e=jpTYTh",
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/EVeGxvEVD_lImKDr4qN1Xi8BvRcpwh9skjygMqnrD2haNw?e=3jFApp",
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/ERNIdAesE0NFkVfJ2S4DSLsBQPiSuPRMgm8K1uPegH-HSQ?e=RbtNAH",
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/EQljDJm7hgVAkpsqnxtCRnIByX5eNyII9uhtPWRop_beqg?e=Fgcx1M",
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/Eboodz5zjHdHquVfonKiISAB2rsOlrfzgWR1HR23AjLmdw?e=fWhcLh",
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/EaS03PhfIgFDhvJiEkIbLhwByUpqqUwU4us4lqYecngtzw?e=KblA9h",
    "https://itsqmet.sharepoint.com/:v:/s/VINCULACINABR-SEP2025429/EX3ov5EOWGZGg_ECvC7KtQwBRwhNL594GSjzSjZzKRFdDg?e=e6hWTV"
]

fechas = [
    datetime(2025, 6, 3),
    datetime(2025, 6, 27),
    datetime(2025, 7, 4),
    datetime(2025, 7, 6),
    datetime(2025, 7, 9),
    datetime(2025, 7, 21),
    datetime(2025, 7, 30)
]

hora_inicio = [
    "21:00",
    "21:00",
    "20:30",
    "19:30",
    "21:00",
    "20:30",
    "20:30"
]

hora_fin = [
    "21:30",
    "21:30",
    "21:00",
    "20:00",
    "21:30",
    "21:00",
    "21:00"
]

actividades = [
    "Introducción al proyecto y presentación de actividades",
    "Desarrollo del capítulo 4 y análisis del producto entregable",
    "Desarrollo del capítulo 4 y avance del producto entregable",
    "Desarrollo del capítulo 5 y avance del producto entregable",
    "Desarrollo del capítulo 5 y avance del producto entregable",
    "Desarrollo del capítulo 5 y avance del producto entregable",
    "Desarrollo del capítulo 5 y avance del producto entregable"
]

descripciones_objetivos = obtener_actividad_objetivo(actividades)

start_row = 5
for i in range(len(reuniones)):
    ws.cell(row=start_row + i*8, column=2, value=descripciones_objetivos[i].objetivo)
    ws.cell(row=start_row + i*8 + 3, column=1, value=fechas[i])
    ws.cell(row=start_row + i*8 + 3, column=2, value=hora_inicio[i])
    ws.cell(row=start_row + i*8 + 3, column=3, value=hora_fin[i])
    ws.cell(row=start_row + i*8 + 3, column=4, value=obtener_duracion(hora_inicio[i], hora_fin[i]))
    ws.cell(row=start_row + i*8 + 3, column=5, value=descripciones_objetivos[i].actividad)
    ws.cell(row=start_row + i*8 + 3, column=6, value=reuniones[i])

output_path = "Vinculacion_Generado.xlsx"
wb.save(output_path)

